#!/usr/bin/env python3
"""Import Narathiwat vehicle-alert workbook and link PowerPoint photos.

Dry run never writes. Apply mode uses SUPABASE_SERVICE_ROLE_KEY from the
environment, supplied only for the process by the PowerShell wrapper.
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
import mimetypes
import os
import re
import sys
import urllib.parse
import urllib.error
import urllib.request
import zipfile
from collections import defaultdict
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook

DEFAULT_PROJECT_URL = "https://rbahodbdbxfvftfxeipe.supabase.co"
TEXT_NS = "{http://schemas.openxmlformats.org/drawingml/2006/main}t"
REL_NS = "{http://schemas.openxmlformats.org/package/2006/relationships}Relationship"
VEHICLE_EVIDENCE_BUCKET = "vehicle-evidence"
ALLOWED_MEDIA_TYPES = {"image/jpeg", "image/png", "image/webp", "image/gif"}


class SupabaseRequestError(RuntimeError):
    """Preserve Storage's response body so import failures are actionable."""

    def __init__(self, status, url, message):
        self.status = status
        self.url = url
        self.message = message
        super().__init__(f"Supabase HTTP {status}: {message}")


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def clean_or_none(value):
    value = clean(value)
    return value or None


def normal_plate(value):
    return "".join(ch for ch in clean(value).upper() if ch.isalnum())


def number(value):
    try:
        return float(value) if value not in (None, "") else None
    except (TypeError, ValueError):
        return None


def date_text(value):
    if not value:
        return None
    if isinstance(value, dt.datetime):
        value = value.date()
    if isinstance(value, dt.date):
        if value.year > 2400:
            value = value.replace(year=value.year - 543)
        return value.isoformat()
    value = clean(value)
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            parsed = dt.datetime.strptime(value[:10], fmt).date()
            if parsed.year > 2400:
                parsed = parsed.replace(year=parsed.year - 543)
            return parsed.isoformat()
        except ValueError:
            pass
    return value


def time_text(value):
    """Return a valid HH:MM:SS value from Excel's inconsistent time cells."""
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        value = value.time()
    if isinstance(value, dt.time):
        return value.replace(microsecond=0).isoformat()

    raw = clean(value)
    if not raw or raw in {"-", "–", "—", "ไม่ระบุ"}:
        return None

    # Examples in source data: "00.01 -", "08:35 น.", "8.35".
    match = re.search(
        r"(?<!\d)([01]?\d|2[0-3])\s*[:.]\s*([0-5]\d)(?:\s*[:.]\s*([0-5]\d))?",
        raw,
    )
    if not match:
        return None
    return f"{int(match.group(1)):02d}:{int(match.group(2)):02d}:{int(match.group(3) or 0):02d}"


def iso_datetime(date_value, time_value):
    normalized_date = date_text(date_value)
    if not normalized_date:
        return None
    try:
        date_part = normalized_date.split("T", 1)[0]
        parsed_date = dt.date.fromisoformat(date_part)
    except ValueError:
        return None
    return f"{parsed_date.isoformat()}T{time_text(time_value) or '00:00:00'}+07:00"


def read_workbook(path: Path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["analyze2"] if "analyze2" in workbook.sheetnames else workbook.active
    iterator = sheet.iter_rows(values_only=True)
    headers = [clean(value) for value in next(iterator)]
    index = {header: position for position, header in enumerate(headers)}

    def get(row, header):
        position = index.get(header)
        return row[position] if position is not None and position < len(row) else None

    records = []
    for fallback_sequence, row in enumerate(iterator, start=1):
        plate = clean_or_none(get(row, "ทะเบียน"))
        if not plate:
            continue
        source_sequence = get(row, "ลำดับ")
        try:
            source_sequence = int(source_sequence)
        except (TypeError, ValueError):
            source_sequence = fallback_sequence
        alert_type = clean_or_none(get(row, "Texttb")) or clean_or_none(get(row, "สถานะ")) or "รถแจ้งเตือน"
        current_status = clean_or_none(get(row, "สถานะ")) or "ใช้งาน"
        if "ได้คืน" in alert_type or "ได้คืน" in current_status:
            current_status = "พบแล้ว"
        photo_refs = [clean_or_none(get(row, header)) for header in ("รย3", "รย4")]
        photo_refs = [reference for reference in photo_refs if reference]
        record = {
            "source_record_key": f"vehicle-alert:{source_sequence}",
            "source_sequence": source_sequence,
            "plate_number": plate,
            "province": clean_or_none(get(row, "ภ.จว.")),
            "vehicle_type": clean_or_none(get(row, "ประเภท")) or "ไม่ระบุ",
            "vehicle_color": clean_or_none(get(row, "สี")),
            "alert_type": alert_type,
            "status": current_status,
            "reported_at": iso_datetime(get(row, "รายงานหาย"), get(row, "เวลาหาย")),
            "last_seen_at": None,
            "last_location": clean_or_none(get(row, "สถานที่หาย")),
            "latitude": number(get(row, "lat")),
            "longitude": number(get(row, "long")),
            "notes": clean_or_none(get(row, "สาเหตุ")),
            "vehicle_brand": clean_or_none(get(row, "ยี่ห้อ")),
            "vehicle_model": clean_or_none(get(row, "รุ่น")),
            "vehicle_year": int(get(row, "car year")) if str(get(row, "car year") or "").isdigit() else None,
            "engine_number": clean_or_none(get(row, "เลขเครื่อง")),
            "chassis_number": clean_or_none(get(row, "เลขตัวรถ")),
            "watch_status": clean_or_none(get(row, "การเฝ้าระวัง")),
            "police_station": clean_or_none(get(row, "สภ.")),
            "police_province": clean_or_none(get(row, "ภ.จว.")),
            "incident_cause": clean_or_none(get(row, "สาเหตุ")),
            "place_type": clean_or_none(get(row, "ประเภทสถานที่")),
            "recovered_at": date_text(get(row, "รายงานได้คืน")),
            "case_status": clean_or_none(get(row, "สถานะ")),
            "security_flag": clean_or_none(get(row, "ความมั่นคง")),
            "ry3_ref": clean_or_none(get(row, "รย3")),
            "ry4_ref": clean_or_none(get(row, "รย4")),
            "lost_time": clean_or_none(get(row, "เวลาหาย")),
            "religion": clean_or_none(get(row, "ศาสนา")),
            "source_photo_refs": photo_refs,
            "source_file": path.name,
            "imported_at": dt.datetime.now(dt.timezone.utc).isoformat(),
        }
        records.append(record)
    return records


def read_workbook_positional(path: Path):
    """Read the authoritative worksheet by fixed source column position.

    Thai headers from this legacy workbook vary by Excel encoding, while its
    source-column positions are stable.  Using positions keeps imports safe.
    """
    columns = {
        "sequence": 0, "alert_type": 1, "case_status": 2,
        "vehicle_type": 3, "watch_status": 4, "plate": 5,
        "brand": 6, "model": 7, "year": 8, "color": 9,
        "engine": 10, "chassis": 11, "reported_at": 12,
        "recovered_at": 13, "station": 14, "province": 15,
        "cause": 16, "place_type": 17, "location": 18,
        "lat": 19, "lng": 20, "security": 21, "ry3": 22,
        "ry4": 23, "lost_time": 24, "religion": 25,
    }
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["analyze2"] if "analyze2" in workbook.sheetnames else workbook.active
    iterator = sheet.iter_rows(values_only=True)
    next(iterator, None)

    def get(row, key):
        index = columns[key]
        return row[index] if index < len(row) else None

    records = []
    for fallback_sequence, row in enumerate(iterator, start=1):
        plate = clean_or_none(get(row, "plate"))
        if not plate:
            continue
        try:
            source_sequence = int(get(row, "sequence"))
        except (TypeError, ValueError):
            source_sequence = fallback_sequence
        source_refs = [clean_or_none(get(row, key)) for key in ("ry3", "ry4")]
        source_refs = [item for item in source_refs if item]
        case_status = clean_or_none(get(row, "case_status")) or "active"
        alert_type = clean_or_none(get(row, "alert_type")) or "vehicle-alert"
        try:
            year = int(get(row, "year")) if get(row, "year") not in (None, "") else None
        except (TypeError, ValueError):
            year = None
        records.append({
            "source_record_key": f"vehicle-alert:{source_sequence}",
            "source_sequence": source_sequence,
            "plate_number": plate,
            "province": clean_or_none(get(row, "province")),
            "vehicle_type": clean_or_none(get(row, "vehicle_type")) or "unspecified",
            "vehicle_color": clean_or_none(get(row, "color")),
            "alert_type": alert_type,
            "status": case_status,
            "reported_at": iso_datetime(get(row, "reported_at"), get(row, "lost_time")),
            "last_seen_at": None,
            "last_location": clean_or_none(get(row, "location")),
            "latitude": number(get(row, "lat")),
            "longitude": number(get(row, "lng")),
            "notes": clean_or_none(get(row, "cause")),
            "vehicle_brand": clean_or_none(get(row, "brand")),
            "vehicle_model": clean_or_none(get(row, "model")),
            "vehicle_year": year,
            "engine_number": clean_or_none(get(row, "engine")),
            "chassis_number": clean_or_none(get(row, "chassis")),
            "watch_status": clean_or_none(get(row, "watch_status")),
            "police_station": clean_or_none(get(row, "station")),
            "police_province": clean_or_none(get(row, "province")),
            "incident_cause": clean_or_none(get(row, "cause")),
            "place_type": clean_or_none(get(row, "place_type")),
            "recovered_at": date_text(get(row, "recovered_at")),
            "case_status": case_status,
            "security_flag": clean_or_none(get(row, "security")),
            "ry3_ref": clean_or_none(get(row, "ry3")),
            "ry4_ref": clean_or_none(get(row, "ry4")),
            "lost_time": clean_or_none(get(row, "lost_time")),
            "religion": clean_or_none(get(row, "religion")),
            "photo_urls": [],
            "source_photo_refs": source_refs,
            "source_file": path.name,
            "imported_at": dt.datetime.now(dt.timezone.utc).isoformat(),
        })
    return records


def powerpoint_media_by_plate(path: Path):
    if not path:
        return {}
    result = defaultdict(list)
    with zipfile.ZipFile(path) as archive:
        names = set(archive.namelist())
        for slide_name in sorted(name for name in names if re.fullmatch(r"ppt/slides/slide\d+\.xml", name)):
            slide_xml = ET.fromstring(archive.read(slide_name))
            text = " ".join(node.text or "" for node in slide_xml.iter(TEXT_NS))
            normalized_text = normal_plate(text)
            rel_name = slide_name.replace("ppt/slides/", "ppt/slides/_rels/") + ".rels"
            if rel_name not in names:
                continue
            rel_xml = ET.fromstring(archive.read(rel_name))
            media = []
            for relationship in rel_xml.findall(REL_NS):
                target = relationship.attrib.get("Target", "")
                if "/media/" not in target:
                    continue
                media_name = "ppt/" + target.replace("../", "")
                if media_name in names:
                    media.append((media_name, archive.read(media_name)))
            if not media:
                continue
            for plate in re.findall(r"[ก-๙A-Z0-9]{1,4}\s*[- ]?\d{1,4}", text.upper()):
                key = normal_plate(plate)
                if key and key in normalized_text:
                    result[key].extend(media)
    return result


def powerpoint_media_by_known_plate(path: Path, plates):
    """Attach slide images to the known plate numbers found in each slide."""
    if not path:
        return {}
    wanted = {plate for plate in plates if len(plate) >= 4}
    result = defaultdict(list)
    with zipfile.ZipFile(path) as archive:
        names = set(archive.namelist())
        for slide_name in sorted(name for name in names if re.fullmatch(r"ppt/slides/slide\d+\.xml", name)):
            slide_xml = ET.fromstring(archive.read(slide_name))
            text = " ".join(node.text or "" for node in slide_xml.iter(TEXT_NS))
            normalized_text = normal_plate(text)
            matched = [plate for plate in wanted if plate in normalized_text]
            if not matched:
                continue
            rel_name = slide_name.replace("ppt/slides/", "ppt/slides/_rels/") + ".rels"
            if rel_name not in names:
                continue
            rel_xml = ET.fromstring(archive.read(rel_name))
            media = []
            for relationship in rel_xml.findall(REL_NS):
                target = relationship.attrib.get("Target", "")
                if "/media/" not in target:
                    continue
                media_name = "ppt/" + target.replace("../", "")
                if media_name in names:
                    media.append((media_name, archive.read(media_name)))
            for plate in matched:
                result[plate].extend(media)
    return result


def request(url, method, key, body=None, content_type="application/json", extra_headers=None):
    headers = {"apikey": key, "Authorization": f"Bearer {key}"}
    if content_type:
        headers["Content-Type"] = content_type
    if extra_headers:
        headers.update(extra_headers)
    data = body if isinstance(body, bytes) else (json.dumps(body, ensure_ascii=False).encode("utf-8") if body is not None else None)
    req = urllib.request.Request(url, data=data, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req, timeout=90) as response:
            return response.read()
    except urllib.error.HTTPError as error:
        message = error.read().decode("utf-8", errors="replace").strip()
        raise SupabaseRequestError(error.code, url, message or error.reason) from error


def ensure_vehicle_evidence_bucket(project_url, key):
    """Create the private Storage bucket when the migration was not run yet."""
    base = project_url.rstrip("/") + "/storage/v1/bucket"
    try:
        request(f"{base}/{VEHICLE_EVIDENCE_BUCKET}", "GET", key)
        return
    except SupabaseRequestError as error:
        # Storage returns either 400 or 404 for a non-existent bucket,
        # depending on the Storage API version.
        if error.status not in {400, 404}:
            raise

    bucket = {
        "id": VEHICLE_EVIDENCE_BUCKET,
        "name": VEHICLE_EVIDENCE_BUCKET,
        "public": False,
        "file_size_limit": 10 * 1024 * 1024,
        "allowed_mime_types": sorted(ALLOWED_MEDIA_TYPES),
    }
    try:
        request(base, "POST", key, bucket)
        print("Created Storage bucket: vehicle-evidence", file=sys.stderr)
    except SupabaseRequestError as error:
        # Another import may have created it between the GET and POST.
        if error.status not in {400, 409}:
            raise


def upload_media(project_url, key, media_by_plate, records):
    uploaded = 0
    skipped = 0
    plate_paths = defaultdict(list)
    try:
        ensure_vehicle_evidence_bucket(project_url, key)
    except SupabaseRequestError as error:
        print(f"Warning: unable to prepare Storage bucket; vehicle records will still import. {error}", file=sys.stderr)
        return uploaded, len(media_by_plate)
    unique_media = {}
    for plate, media_items in media_by_plate.items():
        for media_name, content in media_items:
            digest = f"{plate}/{Path(media_name).name}"
            unique_media[digest] = (media_name, content)
    for storage_path, (media_name, content) in unique_media.items():
        mime = mimetypes.guess_type(media_name)[0] or "application/octet-stream"
        if mime not in ALLOWED_MEDIA_TYPES:
            skipped += 1
            continue
        endpoint = project_url.rstrip("/") + "/storage/v1/object/" + VEHICLE_EVIDENCE_BUCKET + "/" + urllib.parse.quote("vehicle-alerts/" + storage_path, safe="/")
        try:
            request(endpoint, "POST", key, content, mime, {"x-upsert": "true"})
        except SupabaseRequestError as error:
            skipped += 1
            print(f"Warning: skipped media {media_name} ({error})", file=sys.stderr)
            continue
        plate_paths[storage_path.split("/", 1)[0]].append(storage_path)
        uploaded += 1
    for record in records:
        record["photo_urls"] = [
            {"bucket": "vehicle-evidence", "path": "vehicle-alerts/" + storage_path}
            for storage_path in plate_paths.get(normal_plate(record["plate_number"]), [])
        ]
    return uploaded, skipped


def upsert_records(project_url, key, records):
    """Save imported rows without losing an actionable Supabase error.

    Older deployments used a partial unique index for ``source_record_key``.
    PostgREST cannot use a partial index with ``on_conflict``, which results in
    an opaque HTTP 400.  Try the fast batch upsert first, then fall back to an
    update-or-insert request per source row.  The fallback works with both the
    legacy index and the newer unique constraint migration.
    """
    endpoint = project_url.rstrip("/") + "/rest/v1/vehicle_alerts?on_conflict=source_record_key"
    base_endpoint = project_url.rstrip("/") + "/rest/v1/vehicle_alerts"
    for start in range(0, len(records), 100):
        batch = records[start:start + 100]
        try:
            request(endpoint, "POST", key, batch, extra_headers={"Prefer": "resolution=merge-duplicates,return=minimal"})
            continue
        except SupabaseRequestError as error:
            if error.status not in {400, 409}:
                raise
            print(
                "Batch upsert was rejected; using compatible row-by-row save. "
                f"Supabase said: {error.message}",
                file=sys.stderr,
            )

        for record in batch:
            source_key = clean(record.get("source_record_key"))
            try:
                if source_key:
                    filter_url = base_endpoint + "?source_record_key=eq." + urllib.parse.quote(source_key, safe="")
                    updated = request(filter_url, "PATCH", key, record, extra_headers={"Prefer": "return=representation"})
                    if json.loads(updated.decode("utf-8") or "[]"):
                        continue
                request(base_endpoint, "POST", key, [record], extra_headers={"Prefer": "return=minimal"})
            except SupabaseRequestError as row_error:
                label = clean(record.get("plate_number")) or source_key or "unknown row"
                raise RuntimeError(
                    f"Unable to save vehicle record '{label}': {row_error.message}"
                ) from row_error


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", required=True, type=Path)
    parser.add_argument("--pptx", type=Path)
    parser.add_argument("--project-url", default=DEFAULT_PROJECT_URL)
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()
    if not args.xlsx.is_file():
        raise SystemExit(f"Workbook not found: {args.xlsx}")
    if args.pptx and not args.pptx.is_file():
        raise SystemExit(f"PowerPoint not found: {args.pptx}")
    # The source workbook has a fixed, verified column order.  Do not rely on
    # rendered Thai headers because spreadsheet encodings differ by platform.
    records = read_workbook_positional(args.xlsx)
    media = powerpoint_media_by_known_plate(
        args.pptx,
        {normal_plate(row["plate_number"]) for row in records},
    ) if args.pptx else {}
    matched_records = sum(1 for row in records if normal_plate(row["plate_number"]) in media)
    summary = {"source_rows": len(records), "pptx_plate_groups": len(media), "records_with_photos": matched_records, "mode": "apply" if args.apply else "dry-run"}
    if not args.apply:
        print(json.dumps(summary, ensure_ascii=False))
        return
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "").strip()
    if not key:
        raise SystemExit("Set SUPABASE_SERVICE_ROLE_KEY temporarily before running this importer.")
    uploaded, skipped = upload_media(args.project_url, key, media, records)
    upsert_records(args.project_url, key, records)
    summary["uploaded_media"] = uploaded
    summary["skipped_media"] = skipped
    print(json.dumps(summary, ensure_ascii=False))


if __name__ == "__main__":
    main()
